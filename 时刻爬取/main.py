import os
import re
import requests
import pandas as pd
from io import BytesIO
from PIL import Image
import pytesseract
import platform
from datetime import datetime, timedelta
import cv2
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

# 1. Tesseract 路径配置
if platform.system() == "Windows":
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

# 2. 目录配置
base_dir = "."
image_dir = os.path.join(base_dir, "data", "raw_images")
excel_dir = os.path.join(base_dir, "data", "metadata")
excel_path = os.path.join(excel_dir, "表.xlsx")

os.makedirs(image_dir, exist_ok=True)
os.makedirs(excel_dir, exist_ok=True)

# 3. 读取现有表格，防止重复抓取
existing_times = set()
if os.path.exists(excel_path):
    try:
        df = pd.read_excel(excel_path)
        if '时间' in df.columns:
            existing_times = set(df['时间'].astype(str).tolist())
    except Exception as e:
        print(f"读取旧表格失败: {e}")

def get_real_time_from_image(img_bytes):
    """全天候多模式 OCR 识别 + 严格日期校验"""
    try:
        img = Image.open(BytesIO(img_bytes))
        width, height = img.size
        crop_box = (width - 600, 0, width, 150)
        cropped_img = img.crop(crop_box)

        img_array = np.array(cropped_img)
        if len(img_array.shape) == 3:
            gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
        else:
            gray = img_array
            
        gray = cv2.resize(gray, None, fx=2, fy=2, interpolation=cv2.INTER_CUBIC)

        _, thresh_night = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)
        thresh_day = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 31, 2)
        _, thresh_bright = cv2.threshold(gray, 220, 255, cv2.THRESH_BINARY_INV)
        
        filters = [
            ("黑夜模式", thresh_night),
            ("白天模式", thresh_day),
            ("强光模式", thresh_bright),
            ("原图模式", gray)
        ]

        custom_config = r'--oem 3 --psm 6'

        for mode_name, processed_img in filters:
            text = pytesseract.image_to_string(processed_img, lang='eng', config=custom_config)
            
            time_match = re.search(r'(\d{2}:\d{2}:\d{2})', text)
            if time_match:
                hour_minute_second = time_match.group(1)
                date_match = re.search(r'(\d{4}[-/]\d{1,2}[-/]\d{1,2}|\d{1,2}[-/]\d{1,2})', text)
                bjt_now = datetime.utcnow() + timedelta(hours=8)
                
                if date_match:
                    raw_date = date_match.group(1).replace('/', '-')
                    if len(raw_date) <= 5:
                        raw_date = f"{bjt_now.year}-{raw_date}"
                else:
                    raw_date = bjt_now.strftime('%Y-%m-%d')

                final_str = f"{raw_date} {hour_minute_second}"
                
                # 🛡️ 终极安全校验：测试这个日期是否符合人类常识
                try:
                    datetime.strptime(final_str, '%Y-%m-%d %H:%M:%S')
                    print(f"👁️ [{mode_name}] 成功识别时间: {final_str}")
                    return final_str
                except ValueError:
                    # 如果报错（比如19月48日），强行用今天的系统日期 + 识别出的时分秒
                    safe_date = bjt_now.strftime('%Y-%m-%d')
                    safe_str = f"{safe_date} {hour_minute_second}"
                    print(f"⚠️ 发现离谱日期，已自动修正为: {safe_str}")
                    return safe_str
                
        print("⚠️ 所有视觉模式均未能识别出时间文字。")
    except Exception as e:
        print(f"OCR 识别出错: {e}")
    return None

def get_weather_data():
    """调用和风天气 API 获取当前气象数据"""
    api_key = os.environ.get("QWEATHER_API_KEY") 
    if not api_key:
        print("⚠️ 未找到 QWEATHER_API_KEY 环境变量，跳过气象数据获取。")
        return None
        
    url = "https://p46r6yvpw2.re.qweatherapi.com/v7/weather/now"
    params = {
        "location": "116.380,39.977", # 🎯 已更新为你的精确经纬度
        "key": api_key,
        "lang": "zh",
        "unit": "m"
    }
    
    try:
        response = requests.get(url, params=params, timeout=10)
        data = response.json()
        if str(data.get("code")) == "200":
            vis = data["now"]["vis"]
            print(f"✅ 气象数据获取成功！能见度: {vis}km, 观测时间: {data['now']['obsTime']}")
            return data["now"]
        else:
            print(f"❌ 气象 API 返回错误码: {data.get('code')}")
    except Exception as e:
        print(f"🚨 气象 API 请求异常: {e}")
    return None

# 4. 爬取目标
targets = [
    {"url": "http://view.iap.ac.cn:8080/imageview/northeast.jpg", "direction": "northeast"},
    {"url": "http://view.iap.ac.cn:8080/imageview/southwest.jpg", "direction": "southwest"}
]

new_data = []

# 🌟 在抓图前，先拉取一次气象数据
print("\n正在获取实时气象数据...")
weather_info = get_weather_data()
vis_time = weather_info.get("obsTime", "") if weather_info else ""
vis_val = weather_info.get("vis", "") if weather_info else ""
weather_text = weather_info.get("text", "") if weather_info else ""
temp = weather_info.get("temp", "") if weather_info else ""
remark_info = f"{weather_text}, {temp}℃" if weather_info else ""

for target in targets:
    url = target["url"]
    direction = target["direction"]
    print(f"\n正在访问网页: {url}")

    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        img_bytes = response.content

        print("正在启动全天候识别引擎...")
        extracted_time = get_real_time_from_image(img_bytes)
        bjt_now = datetime.utcnow() + timedelta(hours=8)

        if extracted_time:
            final_time = extracted_time
            time_source = 'OCR'
            is_valid = 1
        else:
            final_time = bjt_now.strftime('%Y-%m-%d %H:%M:%S')
            time_source = 'System'
            is_valid = 0

        if final_time in existing_times:
            print(f"🛑 时间 {final_time} 已存在，跳过。")
            continue

        # --- 开始：时间解析与误差计算 ---
        try:
            final_dt = datetime.strptime(final_time, '%Y-%m-%d %H:%M:%S')
            is_daytime = 1 if 6 <= final_dt.hour < 18 else 0
            
            time_diff_min = ''
            if vis_time:
                # 截取和风天气时间前16位 (例如 "2026-04-21T20:15")
                api_dt_str = vis_time[:16].replace('T', ' ')
                api_dt = datetime.strptime(api_dt_str, '%Y-%m-%d %H:%M')
                
                # 计算绝对误差
                diff_seconds = abs((api_dt - final_dt).total_seconds())
                time_diff_min = round(diff_seconds / 60)
                
        except Exception as e:
            print(f"⚠️ 时间解析或误差计算失败: {e}")
            time_diff_min = ''
            is_daytime = ''
        # --- 结束：时间解析与误差计算 ---

        # 🛑 核心拦截器：误差大于 30 分钟直接丢弃！
        if isinstance(time_diff_min, int) and time_diff_min > 30:
            print(f"🗑️ 剔除劣质数据：图片时间 [{final_time}] 与气象时间相差 {time_diff_min} 分钟，超过 30 分钟阈值，已放弃保存。")
            continue

        # 保存图片
        safe_time_str = final_time.replace(':', '-')
        image_filename = f"{safe_time_str}_{direction}.jpg"
        image_path = os.path.join(image_dir, image_filename)

        with open(image_path, 'wb') as f:
            f.write(img_bytes)
        print(f"✅ 保存图片: {image_filename}")

        # 记录元数据
        new_data.append({
            '时间': final_time, '方向': direction, '图片路径': image_path, '图片名': image_filename,
            'time_source': time_source, 'is_valid': is_valid, 'is_daytime': is_daytime,
            'visibility_time': vis_time, 'visibility': vis_val, 'time_diff_min': time_diff_min,
            'label': '', 'remark': remark_info
        })

    except Exception as e:
        print(f"抓取 {direction} 失败: {e}")

# 5. 更新 Excel 表格
if new_data:
    headers = [
        '时间', '方向', '图片路径', '图片名', 'time_source', 'is_valid', 
        'is_daytime', 'visibility_time', 'visibility', 'time_diff_min', 'label', 'remark'
    ]
    
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "爬虫数据"
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        wb.save(excel_path)

    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        for row in new_data:
            ws.append([row[h] for h in headers])
            for cell in ws[ws.max_row]:
                cell.alignment = Alignment(horizontal='left')
        wb.save(excel_path)
        print(f"\n✅ 数据已追加至: {excel_path}")
    except Exception as e:
        print(f"保存表格失败: {e}")
else:
    print("没有新数据，表格未更新。")
